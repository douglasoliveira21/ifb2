"""Cliente para a planilha "Taxas de Rendimento Escolar" — INEP, publicada
anualmente em https://www.gov.br/inep/pt-br/acesso-a-informacao/dados-abertos/
indicadores-educacionais/taxas-de-rendimento.

Diferente do IDEB (uma planilha só, com todos os anos em colunas), esta
fonte publica **um arquivo por ano** — a URL muda a cada edição e cada
arquivo só tem o ano corrente, não o histórico. Por isso, por ora, o IFB
só sincroniza o ano mais recente disponível (2025); séries históricas
exigiriam baixar um arquivo por ano (2012–2024 têm arquivos próprios no
mesmo padrão de URL) — não implementado ainda.

Estrutura da planilha (aba única, cabeçalho em duas linhas nas linhas
6-9): colunas fixas (Ano, Unidade Geográfica, Localização, Dependência
Administrativa) seguidas de blocos de Taxa de Aprovação (prefixo "1_"),
Taxa de Reprovação ("2_"), Taxa de Abandono ("3_"), cada bloco repetindo
Ensino Fundamental (Total, Anos Iniciais, Anos Finais, 1º-9º ano) e
Ensino Médio (Total, 1ª-4ª série, Não-Seriado). O IFB lê só a coluna
"Taxa de Abandono — Ensino Médio — Total" (índice fixo, ver
`_ABANDONO_MEDIO_COL`), filtrando Localização="Total" e Dependência
Administrativa="Total" (todas as redes, urbano+rural combinados).

Conferido: Brasil 2025 = 2,2%; São Paulo 2025 = 2,8% — na faixa
já documentada para abandono no Ensino Médio no Brasil.
"""
import io
import time
import zipfile
from dataclasses import dataclass
from datetime import date

import httpx
import openpyxl

from app.sync.bcb_client import SeriesPoint
from app.sync.inep_client import _ssl_context

REQUEST_HEADERS = {
    "User-Agent": "IFB-Sync/1.0 (+https://github.com/douglasoliveira21/ifb2)",
}
MAX_ATTEMPTS = 3

TX_REND_ZIP_URL = (
    "https://download.inep.gov.br/informacoes_estatisticas/indicadores_educacionais/"
    "2025/tx_rend_brasil_regioes_ufs_2025.zip"
)
TX_REND_YEAR = 2025

_HEADER_ROW = 9
_DATA_START_ROW = 10
_UNIDGEO_COL = 2  # 1-indexado
_LOCALIZACAO_COL = 3
_DEPENDENCIA_COL = 4
_ABANDONO_MEDIO_COL = 53  # "3_CAT_MED" — Taxa de Abandono, Ensino Médio, Total (ver docstring)

_STATE_NAME_TO_UF = {
    "Acre": "AC", "Alagoas": "AL", "Amapá": "AP", "Amazonas": "AM", "Bahia": "BA",
    "Ceará": "CE", "Distrito Federal": "DF", "Espírito Santo": "ES", "Goiás": "GO",
    "Maranhão": "MA", "Mato Grosso": "MT", "Mato Grosso do Sul": "MS", "Minas Gerais": "MG",
    "Pará": "PA", "Paraíba": "PB", "Paraná": "PR", "Pernambuco": "PE", "Piauí": "PI",
    "Rio de Janeiro": "RJ", "Rio Grande do Norte": "RN", "Rio Grande do Sul": "RS",
    "Rondônia": "RO", "Roraima": "RR", "Santa Catarina": "SC", "São Paulo": "SP",
    "Sergipe": "SE", "Tocantins": "TO",
}


@dataclass(frozen=True)
class TaxaRendimentoSpec:
    url: str = TX_REND_ZIP_URL
    year: int = TX_REND_YEAR


DEFAULT_SPEC = TaxaRendimentoSpec()


def _download_workbook(spec: TaxaRendimentoSpec, *, timeout: float):
    ctx = _ssl_context()
    content = None
    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            response = httpx.get(spec.url, headers=REQUEST_HEADERS, timeout=timeout, verify=ctx, follow_redirects=True)
            response.raise_for_status()
            content = response.content
            break
        except httpx.TransportError:
            if attempt < MAX_ATTEMPTS:
                time.sleep(2 * attempt)
                continue
            raise
    zf = zipfile.ZipFile(io.BytesIO(content))
    xlsx_name = next(name for name in zf.namelist() if name.endswith(".xlsx"))
    with zf.open(xlsx_name) as f:
        xlsx_bytes = f.read()
    return openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)


def fetch_taxa_abandono_medio_by_state(
    spec: TaxaRendimentoSpec = DEFAULT_SPEC, *, timeout: float = 60.0
) -> dict[str, list[SeriesPoint]]:
    wb = _download_workbook(spec, timeout=timeout)
    ws = wb[wb.sheetnames[0]]

    by_state: dict[str, list[SeriesPoint]] = {}
    for row in ws.iter_rows(min_row=_DATA_START_ROW, values_only=True):
        if row[_LOCALIZACAO_COL - 1] != "Total" or row[_DEPENDENCIA_COL - 1] != "Total":
            continue
        uf = _STATE_NAME_TO_UF.get(row[_UNIDGEO_COL - 1])
        if uf is None:
            continue
        value = row[_ABANDONO_MEDIO_COL - 1]
        if isinstance(value, (int, float)):
            by_state[uf] = [SeriesPoint(reference_date=date(spec.year, 12, 31), value=float(value))]

    return by_state


def fetch_taxa_abandono_medio_brasil(
    spec: TaxaRendimentoSpec = DEFAULT_SPEC, *, timeout: float = 60.0
) -> list[SeriesPoint]:
    wb = _download_workbook(spec, timeout=timeout)
    ws = wb[wb.sheetnames[0]]

    for row in ws.iter_rows(min_row=_DATA_START_ROW, values_only=True):
        if (
            row[_UNIDGEO_COL - 1] == "Brasil"
            and row[_LOCALIZACAO_COL - 1] == "Total"
            and row[_DEPENDENCIA_COL - 1] == "Total"
        ):
            value = row[_ABANDONO_MEDIO_COL - 1]
            if isinstance(value, (int, float)):
                return [SeriesPoint(reference_date=date(spec.year, 12, 31), value=float(value))]
    return []
