"""Cliente para os microdados de divulgação do IDEB — INEP.

Diferente das demais fontes, o INEP não expõe uma API: os resultados do
IDEB são publicados como uma planilha (.xlsx dentro de um .zip), uma
edição por vez. Este módulo baixa o zip, extrai a planilha e lê os
valores observados diretamente das abas "Brasil (...)", sem hardcodar
posições de linha/coluna — a leitura localiza o cabeçalho dinamicamente
para não quebrar se o INEP inserir/remover colunas em edições futuras.

A estrutura foi conferida manualmente contra a série histórica oficial
do IDEB Anos Iniciais (3.8, 4.2, 4.6, 5.0, 5.2, 5.5, 5.8, 5.9, 5.8, 6.0,
6.3 — 2005 a 2025, incluindo a queda de 2021 por causa da pandemia).

**Sobre o certificado TLS**: `download.inep.gov.br` não envia o
certificado intermediário ("RNP ICPEdu GR46 OV TLS CA 2025") durante o
handshake — só o certificado final. Clientes que fazem "AIA chasing"
(buscam o intermediário automaticamente, como o Windows) conseguem
validar mesmo assim; o OpenSSL usado pelo Python em containers Linux não
faz isso, e a conexão falha com `CERTIFICATE_VERIFY_FAILED`. Confirmado
manualmente com `openssl s_client`. A correção não é desativar a
verificação do certificado — é fornecer o intermediário que falta:
`app/sync/certs/rnp_icpedu_gr46_ov_tls_ca_2025.pem` (baixado da própria
URL "CA Issuers" do certificado do INEP, que termina numa raiz pública da
GlobalSign já confiável por padrão).
"""
import functools
import io
import re
import ssl
import time
import zipfile
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import certifi
import httpx
import openpyxl

from app.sync.bcb_client import SeriesPoint

REQUEST_HEADERS = {
    "User-Agent": "IFB-Sync/1.0 (+https://github.com/douglasoliveira21/ifb2)",
}
MAX_ATTEMPTS = 3

_CERTS_DIR = Path(__file__).parent / "certs"
_OBSERVADO_RE = re.compile(r"^VL_OBSERVADO_(\d{4})$")


@functools.lru_cache(maxsize=1)
def _ssl_context() -> ssl.SSLContext:
    """Trust store padrão (certifi) + o intermediário que o INEP não envia."""
    ctx = ssl.create_default_context(cafile=certifi.where())
    for cert_path in sorted(_CERTS_DIR.glob("*.pem")):
        ctx.load_verify_locations(cafile=str(cert_path))
    return ctx


@dataclass(frozen=True)
class IdebSheetSpec:
    """Descreve onde, dentro do zip de divulgação do IDEB, está a série
    nacional de uma etapa de ensino."""

    zip_url: str
    sheet_name: str
    total_row_label_col_b: str = "Total"


def _extract_xlsx_bytes(zip_bytes: bytes) -> bytes:
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        xlsx_names = [name for name in zf.namelist() if name.lower().endswith(".xlsx")]
        if not xlsx_names:
            raise ValueError("nenhum arquivo .xlsx encontrado no zip do IDEB")
        return zf.read(xlsx_names[0])


def _parse_sheet(xlsx_bytes: bytes, sheet_name: str, total_row_label: str) -> list[SeriesPoint]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[sheet_name]

    header_row = None
    year_columns: dict[int, int] = {}
    for row_idx in range(1, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        matches = {
            c + 1: int(m.group(1))
            for c, v in enumerate(row_values)
            if isinstance(v, str) and (m := _OBSERVADO_RE.match(v))
        }
        if matches:
            header_row = row_idx
            year_columns = {year: col for col, year in matches.items()}
            break

    if header_row is None:
        raise ValueError(f"cabeçalho VL_OBSERVADO_* não encontrado na aba '{sheet_name}'")

    total_row = None
    for row_idx in range(header_row + 1, ws.max_row + 1):
        if ws.cell(row=row_idx, column=2).value == total_row_label:
            total_row = row_idx
            break

    if total_row is None:
        raise ValueError(f"linha '{total_row_label}' não encontrada na aba '{sheet_name}'")

    points: list[SeriesPoint] = []
    for year, col in sorted(year_columns.items()):
        value = ws.cell(row=total_row, column=col).value
        if value is None or not isinstance(value, (int, float)):
            continue
        points.append(SeriesPoint(reference_date=date(year, 1, 1), value=float(value)))

    points.sort(key=lambda p: p.reference_date)
    return points


def _download_zip(url: str, *, timeout: float) -> bytes:
    """O servidor de download do INEP é instável sob uso automatizado —
    reseta a conexão ou falha o handshake TLS em parte das tentativas,
    mesmo quando o arquivo está disponível (confirmado manualmente: a
    mesma URL falha na 1ª tentativa e funciona logo em seguida). Por
    isso, algumas tentativas com espera entre elas antes de desistir."""
    last_exc: Exception | None = None
    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            response = httpx.get(
                url,
                headers=REQUEST_HEADERS,
                timeout=timeout,
                follow_redirects=True,
                verify=_ssl_context(),
            )
            response.raise_for_status()
            return response.content
        except (httpx.TransportError, httpx.HTTPStatusError) as exc:
            last_exc = exc
            if attempt < MAX_ATTEMPTS:
                time.sleep(3 * attempt)
    assert last_exc is not None
    raise last_exc


@functools.lru_cache(maxsize=4)
def _cached_xlsx_bytes(zip_url: str, *, timeout: float) -> bytes:
    """As três etapas do IDEB (Anos Iniciais/Finais/EM) vêm do mesmo zip —
    baixar uma vez por execução em vez de uma vez por indicador reduz o
    número de requisições ao servidor instável do INEP em 3x."""
    zip_bytes = _download_zip(zip_url, timeout=timeout)
    return _extract_xlsx_bytes(zip_bytes)


def fetch_ideb_series(spec: IdebSheetSpec, *, timeout: float = 60.0) -> list[SeriesPoint]:
    """Baixa (ou reaproveita, se já baixado nesta execução) o zip de
    divulgação do IDEB e extrai a série nacional (linha 'Total', todas as
    redes) da aba indicada."""
    xlsx_bytes = _cached_xlsx_bytes(spec.zip_url, timeout=timeout)
    return _parse_sheet(xlsx_bytes, spec.sheet_name, spec.total_row_label_col_b)
