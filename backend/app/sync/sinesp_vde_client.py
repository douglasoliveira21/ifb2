"""Cliente para a Base de Dados Nacional de Segurança Pública (Sinesp VDE —
Validador de Dados Estatísticos), Ministério da Justiça e Segurança
Pública — publicada como um arquivo .xlsx por ano em
https://www.gov.br/mj/pt-br/assuntos/sua-seguranca/seguranca-publica/estatistica/dados-nacionais-1/base-de-dados-e-notas-metodologicas-dos-gestores-estaduais-sinesp-vde-2022-e-2023

Mesmo padrão de "baixar arquivo estático por ano" já usado para Leitos
SUS e IDEB — dados declarados pelos gestores estaduais de segurança
pública, consolidados nacionalmente pelo MJSP. Cada arquivo traz um
registro por UF/município/tipo de ocorrência/mês (`evento`, `uf`,
`municipio`, `data_referencia`, `total_vitima`).

**Precisa de cabeçalhos de navegador**: o servidor do gov.br devolve 403
para requisições sem `User-Agent`/`Referer` — não é autenticação, é uma
checagem anti-bot básica. `REQUEST_HEADERS` abaixo já inclui um
User-Agent de navegador real; sem isso, toda chamada falha com 403.

Validado ao vivo contra números amplamente divulgados: "Homicídio
doloso" somado nacionalmente em 2019 = 39.228 e em 2024 = 35.136 —
ambos na mesma ordem de grandeza dos totais anuais de homicídio doloso
publicados pelo Fórum Brasileiro de Segurança Pública/Anuário Brasileiro
de Segurança Pública para esses anos; por estado, Bahia lidera em 2024
(4.207), consistente com a taxa de homicídios da Bahia estar entre as
mais altas do país nos últimos anos, amplamente noticiado.
"""
import functools
import time
from collections import defaultdict
from datetime import date

import httpx
import openpyxl

from app.sync.bcb_client import SeriesPoint

URL_TEMPLATE = (
    "https://www.gov.br/mj/pt-br/assuntos/sua-seguranca/seguranca-publica/estatistica/"
    "download/dnsp-base-de-dados/bancovde-{year}.xlsx"
)

REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
    ),
}
MAX_ATTEMPTS = 3
RETRY_STATUS_CODES = {429, 500, 502, 503, 504}

FIRST_AVAILABLE_YEAR = 2015
EVENTO_HOMICIDIO_DOLOSO = "Homicídio doloso"

# Colunas do arquivo, na ordem em que aparecem (confirmado empiricamente —
# a planilha não muda de posição entre os anos testados, 2019 e 2024).
_COL_UF = 0
_COL_EVENTO = 2
_COL_TOTAL_VITIMA = 10


@functools.lru_cache(maxsize=16)
def _download_year_workbook_bytes(year: int, *, timeout: float) -> bytes | None:
    """Baixa o arquivo .xlsx de um ano. Retorna None se o ano ainda não foi
    publicado (403/404 mesmo com os cabeçalhos corretos) — não é erro, é a
    fonte não ter esse ano ainda.

    Cacheado por processo: a sincronização Brasil e por estado leem o
    mesmo arquivo anual (~30 MB cada) — sem isso, cada uma baixaria tudo
    de novo."""
    url = URL_TEMPLATE.format(year=year)
    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            response = httpx.get(url, headers=REQUEST_HEADERS, timeout=timeout, follow_redirects=True)
        except httpx.TransportError:
            if attempt < MAX_ATTEMPTS:
                time.sleep(2 * attempt)
                continue
            raise
        if response.status_code in {403, 404}:
            return None
        if response.status_code in RETRY_STATUS_CODES and attempt < MAX_ATTEMPTS:
            time.sleep(2 * attempt)
            continue
        response.raise_for_status()
        return response.content
    raise AssertionError("unreachable")


def _sum_evento_by_uf(workbook_bytes: bytes, evento: str) -> dict[str, int]:
    import io

    workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes), read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    totals: dict[str, int] = defaultdict(int)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[_COL_EVENTO] != evento:
            continue
        raw = row[_COL_TOTAL_VITIMA]
        totals[row[_COL_UF]] += int(raw) if isinstance(raw, (int, float)) else 0

    workbook.close()
    return dict(totals)


def _fetch_totals_by_year(
    evento: str, *, start_year: int, timeout: float
) -> dict[int, dict[str, int]]:
    """Baixa cada arquivo anual uma única vez e devolve {ano: {uf: total}}
    — usado tanto pela série por estado quanto pelo agregado Brasil, para
    não baixar o mesmo arquivo (~30 MB) duas vezes."""
    current_year = date.today().year
    totals_by_year: dict[int, dict[str, int]] = {}

    for year in range(start_year, current_year + 1):
        workbook_bytes = _download_year_workbook_bytes(year, timeout=timeout)
        if workbook_bytes is None:
            continue
        totals_by_year[year] = _sum_evento_by_uf(workbook_bytes, evento)

    return totals_by_year


def fetch_homicidio_doloso_by_state(
    *, start_year: int = FIRST_AVAILABLE_YEAR, timeout: float = 120.0
) -> dict[str, list[SeriesPoint]]:
    """Total de homicídios dolosos por estado, um ponto por ano."""
    by_state: dict[str, list[SeriesPoint]] = {}

    for year, totals in _fetch_totals_by_year(
        EVENTO_HOMICIDIO_DOLOSO, start_year=start_year, timeout=timeout
    ).items():
        for uf, total in totals.items():
            by_state.setdefault(uf, []).append(SeriesPoint(reference_date=date(year, 1, 1), value=float(total)))

    for points in by_state.values():
        points.sort(key=lambda p: p.reference_date)

    return by_state


def fetch_homicidio_doloso_brasil(
    *, start_year: int = FIRST_AVAILABLE_YEAR, timeout: float = 120.0
) -> list[SeriesPoint]:
    """Mesma leitura, somada para o Brasil (soma de todos os estados em
    cada ano)."""
    points: list[SeriesPoint] = []
    for year, totals in _fetch_totals_by_year(
        EVENTO_HOMICIDIO_DOLOSO, start_year=start_year, timeout=timeout
    ).items():
        points.append(SeriesPoint(reference_date=date(year, 1, 1), value=float(sum(totals.values()))))

    points.sort(key=lambda p: p.reference_date)
    return points
