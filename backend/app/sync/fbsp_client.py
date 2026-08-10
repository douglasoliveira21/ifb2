"""Cliente para a planilha de dados do Anuário Brasileiro de Segurança
Pública — Fórum Brasileiro de Segurança Pública (FBSP).

**Atenção — única fonte não-governamental do IFB**: diferente de todas as
demais integrações (IBGE, BCB, INPE, INEP, Tesouro Nacional), o FBSP é
uma associação civil sem fins lucrativos, não um órgão público. Ele foi
usado aqui porque nenhuma fonte primária do governo federal para dados
de segurança pública por estado se mostrou utilizável: o SINESP (MJSP)
só publica via `dados.mj.gov.br`, domínio que não resolve mais por DNS
(confirmado — `NXDOMAIN`), e o portal novo (`dados.gov.br`) expõe os
metadados do dataset sem autenticação, mas exige login para baixar
qualquer arquivo (testado: 401 mesmo em sessão de browser autenticada
anonimamente).

O FBSP consolida, valida e publica anualmente os dados que as próprias
Secretarias de Segurança Pública estaduais enviam ao Sinesp — não são
números inventados pelo FBSP, é a mesma fonte primária (estados),
reprocessada e auditada por uma organização de pesquisa amplamente
citada por governo, imprensa e academia. Ainda assim, o IFB marca este
indicador com a fonte real (FBSP, não "Governo Federal") em toda a
interface, para nunca dar a entender que é um número oficial do MJSP.

Os valores da taxa de Mortes Violentas Intencionais (MVI) — que soma
homicídio doloso, latrocínio, lesão corporal seguida de morte e mortes
por intervenção policial — foram conferidos contra a edição 2025 (dados
2023-2024): Brasil 2024 = 20,76 por 100 mil habitantes, batendo com o
número amplamente noticiado na divulgação do próprio Anuário; Bahia
entre os estados com taxa mais alta, consistente com a concentração de
violência letal no Nordeste já documentada por outras fontes.

**Sem série histórica automática**: assim como o IDEB (`inep_client.py`),
o FBSP publica uma planilha por edição anual, sem API — a URL e os
índices de coluna usados aqui (`_MVI_SHEET`, colunas de "Taxa") são
específicos da 19ª edição (2025) e precisarão ser atualizados
manualmente no código a cada nova edição.
"""
import functools
import io
import re
import time
from dataclasses import dataclass
from datetime import date

import httpx
import openpyxl

from app.sync.bcb_client import SeriesPoint

ANUARIO_URL = "https://publicacoes.forumseguranca.org.br/bitstreams/a64087c0-e7f5-419d-b0f3-08c55faca5b1/download"

REQUEST_HEADERS = {
    "User-Agent": "IFB-Sync/1.0 (+https://github.com/douglasoliveira21/ifb2)",
}
MAX_ATTEMPTS = 3
RETRY_STATUS_CODES = {429, 500, 502, 503, 504}

_MVI_SHEET = "T01"
_DATA_START_ROW = 11  # 1-indexed (openpyxl) — primeira linha de dado é "Brasil"
_STATE_NAME_COL = 1
_RATE_COLUMNS_BY_YEAR = {2023: 14, 2024: 15}  # 1-indexado

_STATE_NAME_TO_UF = {
    "Acre": "AC", "Alagoas": "AL", "Amapá": "AP", "Amazonas": "AM", "Bahia": "BA",
    "Ceará": "CE", "Distrito Federal": "DF", "Espírito Santo": "ES", "Goiás": "GO",
    "Maranhão": "MA", "Mato Grosso": "MT", "Mato Grosso do Sul": "MS", "Minas Gerais": "MG",
    "Pará": "PA", "Paraíba": "PB", "Paraná": "PR", "Pernambuco": "PE", "Piauí": "PI",
    "Rio de Janeiro": "RJ", "Rio Grande do Norte": "RN", "Rio Grande do Sul": "RS",
    "Rondônia": "RO", "Roraima": "RR", "Santa Catarina": "SC", "São Paulo": "SP",
    "Sergipe": "SE", "Tocantins": "TO",
}
_FOOTNOTE_SUFFIX_RE = re.compile(r"\s*\(\d+\)\s*$")  # ex: "Minas Gerais (4)" -> "Minas Gerais"


@functools.lru_cache(maxsize=4)
def _download_with_retry(url: str, *, timeout: float) -> bytes:
    # Cacheado por processo: a leitura por estado e a leitura Brasil usam
    # a mesma planilha — evita baixar o arquivo (~1.7 MB) duas vezes.
    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            response = httpx.get(url, headers=REQUEST_HEADERS, timeout=timeout, follow_redirects=True)
        except httpx.TransportError:
            if attempt < MAX_ATTEMPTS:
                time.sleep(2 * attempt)
                continue
            raise
        if response.status_code in RETRY_STATUS_CODES and attempt < MAX_ATTEMPTS:
            time.sleep(2 * attempt)
            continue
        response.raise_for_status()
        return response.content
    raise AssertionError("unreachable")


@dataclass(frozen=True)
class FbspAnuarioSpec:
    url: str = ANUARIO_URL
    sheet_name: str = _MVI_SHEET


DEFAULT_ANUARIO_SPEC = FbspAnuarioSpec()


def fetch_mvi_rate_by_state(
    spec: FbspAnuarioSpec = DEFAULT_ANUARIO_SPEC, *, timeout: float = 60.0
) -> dict[str, list[SeriesPoint]]:
    """Lê a taxa de Mortes Violentas Intencionais (por 100 mil habitantes)
    por estado, para os anos presentes na edição vigente do Anuário
    (atualmente 2023 e 2024). Ignora a linha "Brasil" (é o agregado
    nacional, sincronizado separadamente) e qualquer linha cujo nome não
    bata com uma UF conhecida (notas de rodapé, linhas em branco)."""
    content = _download_with_retry(spec.url, timeout=timeout)
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[spec.sheet_name]

    by_state: dict[str, list[SeriesPoint]] = {}
    for row_idx in range(_DATA_START_ROW, ws.max_row + 1):
        raw_name = ws.cell(row=row_idx, column=_STATE_NAME_COL).value
        if not isinstance(raw_name, str):
            continue
        name = _FOOTNOTE_SUFFIX_RE.sub("", raw_name).strip()
        uf = _STATE_NAME_TO_UF.get(name)
        if uf is None:
            continue

        points: list[SeriesPoint] = []
        for year, col in _RATE_COLUMNS_BY_YEAR.items():
            value = ws.cell(row=row_idx, column=col).value
            if isinstance(value, (int, float)):
                points.append(SeriesPoint(reference_date=date(year, 1, 1), value=float(value)))
        if points:
            by_state[uf] = points

    return by_state


def fetch_mvi_rate_brasil(spec: FbspAnuarioSpec = DEFAULT_ANUARIO_SPEC, *, timeout: float = 60.0) -> list[SeriesPoint]:
    """Mesma leitura, mas só a linha 'Brasil' (agregado nacional)."""
    content = _download_with_retry(spec.url, timeout=timeout)
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[spec.sheet_name]

    for row_idx in range(_DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=row_idx, column=_STATE_NAME_COL).value == "Brasil":
            points = []
            for year, col in _RATE_COLUMNS_BY_YEAR.items():
                value = ws.cell(row=row_idx, column=col).value
                if isinstance(value, (int, float)):
                    points.append(SeriesPoint(reference_date=date(year, 1, 1), value=float(value)))
            return points

    return []
