"""Cliente para a planilha "Base CTB GG.xlsx" (Carga Tributária Bruta do
Governo Geral) — Secretaria do Tesouro Nacional, publicada anualmente em
https://www.tesourotransparente.gov.br/publicacoes/carga-tributaria-do-governo-geral.

Só nível Brasil: a Carga Tributária Bruta soma a arrecadação de todas as
esferas de governo (União, estados, municípios) sobre o PIB nacional —
não é um dado declarado por ente federativo como o SICONFI, então não
existe quebra por estado.

**Sem série histórica automática**: o Tesouro publica um anexo novo a
cada edição (mesmo padrão do IDEB/FBSP já usados no IFB) — a URL do
anexo (`publicacao-anexo/<id>`) muda a cada divulgação e precisa ser
atualizada manualmente no código.

Valores conferidos contra a divulgação oficial da edição 2025: Brasil
2025 = 32,40% do PIB — bate com a manchete "maior valor da série
histórica" amplamente noticiada em abril/2026.
"""
import io
import time
from dataclasses import dataclass
from datetime import date

import httpx
import openpyxl

from app.sync.bcb_client import SeriesPoint

CTB_URL = "https://thot-arquivos.tesouro.gov.br/publicacao-anexo/28024"

REQUEST_HEADERS = {
    "User-Agent": "IFB-Sync/1.0 (+https://github.com/douglasoliveira21/ifb2)",
}
MAX_ATTEMPTS = 3
RETRY_STATUS_CODES = {429, 500, 502, 503, 504}

_SHEET = "Tabela 1"
_HEADER_ROW = 4
_GOVERNO_GERAL_PCT_PIB_ROW = 9  # linha "Governo Geral" do bloco "% do PIB" (o 2º bloco da tabela)


@dataclass(frozen=True)
class CargaTributariaSpec:
    url: str = CTB_URL


DEFAULT_CTB_SPEC = CargaTributariaSpec()


def _download_with_retry(url: str, *, timeout: float) -> bytes:
    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            response = httpx.get(url, headers=REQUEST_HEADERS, timeout=timeout, follow_redirects=True)
        except httpx.TransportError:
            if attempt < MAX_ATTEMPTS:
                time.sleep(2 * attempt)
                continue
            raise
        if response.status_code in RETRY_STATUS_CODES and attempt < MAX_ATTEMPTS:
            time.sleep(2 * attempt)
            continue
        response.raise_for_status()
        return response.content
    raise AssertionError("unreachable")


def fetch_carga_tributaria_brasil(
    spec: CargaTributariaSpec = DEFAULT_CTB_SPEC, *, timeout: float = 60.0
) -> list[SeriesPoint]:
    """Carga Tributária Bruta do Governo Geral (% do PIB), série anual —
    linha "Governo Geral" do segundo bloco da Tabela 1 (valores em % do
    PIB; o primeiro bloco, ignorado aqui, traz os mesmos dados em R$
    milhões). Multiplica por 100: a planilha traz fração (0.324), o IFB
    guarda percentual (32.4) — ver frontend/lib/format.ts."""
    content = _download_with_retry(spec.url, timeout=timeout)
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[_SHEET]

    points: list[SeriesPoint] = []
    for col in range(2, ws.max_column + 1):
        year = ws.cell(row=_HEADER_ROW, column=col).value
        if not isinstance(year, int):
            continue
        value = ws.cell(row=_GOVERNO_GERAL_PCT_PIB_ROW, column=col).value
        if isinstance(value, (int, float)):
            points.append(SeriesPoint(reference_date=date(year, 12, 31), value=float(value) * 100))

    return points
